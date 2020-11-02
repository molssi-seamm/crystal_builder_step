#!/usr/bin/env python3

import json
import openpyxl

subscript = [
    '\N{Subscript Zero}',
    '\N{Subscript One}',
    '\N{Subscript Two}',
    '\N{Subscript Three}',
    '\N{Subscript Four}',
    '\N{Subscript Five}',
    '\N{Subscript Six}',
    '\N{Subscript Seven}',
    '\N{Subscript Eight}',
    '\N{Subscript Nine}',
]


def formula_subscripts(formula):
    result = ''
    for char in formula:
        if char in '0123456789':
            result += subscript[int(char)]
        else:
            result += char
    return result


def run(filename='prototypes.xlsx'):
    data = {}

    wb = openpyxl.load_workbook(filename=filename)
    ws = wb['Sheet1']

    row = 1
    row_max = ws.max_row

    data = {
        'prototype': [],
        'nSpecies': [],
        'nAtoms': [],
        'Pearson symbol': [],
        'Strukturbericht designation': [],
        'AFLOW prototype': [],
        'space group symbol': [],
        'space group number': [],
        'notes': []
    }
    hyperlink = []

    for row in ws.iter_rows(min_row=2, max_row=row_max):
        hyperlink.append(row[0].hyperlink.display)
        for cell, column in zip(row, data.keys()):
            data[column].append(cell.value)

    data['hyperlink'] = hyperlink

    # Now clean up the data
    tmp = []
    # Characters in notes
    for value in data['notes']:
        tmp.append(value.replace('\u00a0', ' '))
    data['notes'] = tmp

    # Subscripts in formulas of prototypes
    tmp = []
    for value in data['prototype']:
        value = formula_subscripts(value)
        print(value)
        tmp.append(value)
    data['prototype'] = tmp

    # doubling of Strukturbericht designations
    tmp = []
    for value in data['Strukturbericht designation']:
        if value is not None:
            length = int(len(value) / 2)
            value = value[0:length]
        tmp.append(value)
    data['Strukturbericht designation'] = tmp

    # doubling of space group symbols
    tmp = []
    for value in data['space group symbol']:
        if value is not None:
            length = int(len(value) / 2)
            value = value[0:length].replace('\u00af', '\N{Combining Overline}')
        print(value)
        tmp.append(value)
    data['space group symbol'] = tmp

    with open('prototypes.json', 'w') as fd:
        json.dump(data, fd, indent=4)


if __name__ == "__main__":  # pragma: no cover
    import os.path
    import urllib.error
    import urllib.parse
    import urllib.request

    # run()

    # Get the data from the json
    with open('prototypes.json', 'r') as fd:
        data = json.load(fd)

    # Fetch the cif file
    for hyperlink in data['hyperlink']:
        url = urllib.parse.urlsplit(hyperlink)
        path = url.path
        root, extension = os.path.splitext(os.path.basename(path))
        filename = root + '.cif'
        new_path = os.path.join('prototypes', filename)
        if os.path.exists(new_path):
            continue
        new_url = urllib.parse.urljoin(hyperlink, 'CIF/' + filename)
        print(f'{new_url} --> {new_path}')
        try:
            with urllib.request.urlopen(new_url) as response:
                with open(new_path, 'w') as fd:
                    fd.write(response.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            print(f'#### Error: {e}')
